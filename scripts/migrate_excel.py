"""
PowerNet Manager — Excel to Supabase migration script.
Run: python scripts/migrate_excel.py
Requires SUPABASE_SERVICE_KEY in .env.local
"""
import os
import re
import datetime
import openpyxl
from supabase import create_client

RAHWALI_FILE  = r'C:\Users\PC\Downloads\APRIL_2026_RAHWALI.xlsx'
GARRISON_FILE = r'C:\Users\PC\Downloads\APRIL 2026 GARRISON.xlsx'

# ── Area definitions ────────────────────────────────────────────────────────

RAHWALI_AREAS = {
    'NEW AIT':      {'code': 'N-AIT', 'name': 'New Alama Iqbal Town', 'type': 'civilian'},
    'AIT':          {'code': 'AIT',   'name': 'Alama Iqbal Town',     'type': 'civilian'},
    'KHUSHI T':     {'code': 'KT',    'name': 'Khushi Town',          'type': 'civilian'},
    'GREEN T':      {'code': 'GT',    'name': 'Green Town',            'type': 'civilian'},
    'MUSLIM T':     {'code': 'MT',    'name': 'Muslim Town',           'type': 'civilian'},
    'SETHI COLONY': {'code': 'SC',    'name': 'Sethi Colony',          'type': 'civilian'},
    'RW SHARQI':    {'code': 'RW',    'name': 'Rahwali Sharqi',        'type': 'civilian'},
    'SHARIF PURA':  {'code': 'SP',    'name': 'Sharif Pura',           'type': 'civilian'},
    'SHARIF FARM':  {'code': 'SF',    'name': 'Sharif Farm',           'type': 'civilian'},
    'MAKI MASJID':  {'code': 'MM',    'name': 'Maki Masjid',           'type': 'civilian'},
    'GHADI SHAHU':  {'code': 'GS',    'name': 'Ghadi Shahu',           'type': 'civilian'},
    'GULAB PURA':   {'code': 'GP',    'name': 'Gulab Pura',            'type': 'civilian'},
    'BILAL TOWN':   {'code': 'BT',    'name': 'Bilal Town',            'type': 'civilian'},
    'DHINGWALI':    {'code': 'DG',    'name': 'Dhingranwali',          'type': 'civilian'},
    'MADINA C':     {'code': 'MC',    'name': 'Madina Colony',         'type': 'civilian'},
    'MAIN BAZAR':   {'code': 'MB',    'name': 'Main Bazar',            'type': 'civilian'},
    'SLP PURA':     {'code': 'SLP',   'name': 'Salmat Pura',           'type': 'civilian'},
    'AMT PURA':     {'code': 'AMT',   'name': 'Amrat Pura',            'type': 'civilian'},
}

GARRISON_AREAS = {
    'ARMY AREA':  {'code': 'AR',      'name': 'Army Area',        'type': 'garrison'},
    'BZR':        {'code': 'BZR',     'name': 'Bazaar',            'type': 'garrison'},
    'ASK 1':      {'code': 'ASK-1',   'name': 'Ask Sector 1',     'type': 'garrison'},
    'ASK 2':      {'code': 'ASK-2',   'name': 'Ask Sector 2',     'type': 'garrison'},
    'DEF 1':      {'code': 'DEF-1',   'name': 'Defence Sector 1', 'type': 'garrison'},
    'DEF 2':      {'code': 'DEF-2',   'name': 'Defence Sector 2', 'type': 'garrison'},
    'GT ROAD':    {'code': 'GT-ROAD', 'name': 'GT Road',           'type': 'civilian'},
    'DC COLONY':  {'code': 'DC',      'name': 'DC Colony',         'type': 'garrison'},
}

# ── Helpers ─────────────────────────────────────────────────────────────────

def load_env():
    env = {}
    env_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env.local')
    if os.path.exists(env_file):
        with open(env_file) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    env[k.strip()] = v.strip()
    return env

def normalize_package(raw):
    if not raw:
        return None
    s = str(raw).strip()
    is_5g = '5G' in s.upper()
    match = re.search(r'(\d+)', s)
    if not match:
        return None
    speed = int(match.group(1))
    return f'{speed} Mbps 5G' if is_5g else f'{speed} Mbps'

def parse_due(due_val):
    """Returns (status, due_amount_or_None)"""
    if due_val is None:
        return ('active', None)
    s = str(due_val).strip().upper()
    if s in ('DC', ''):
        return ('disconnected', None)
    if s == 'TDC':
        return ('tdc', None)
    if s == 'FREE':
        return ('free', None)
    if s.startswith('SHIFT'):
        return ('shifted', None)
    try:
        amount = int(float(due_val))
        return ('active', amount)
    except (ValueError, TypeError):
        return ('active', None)

def parse_date(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        if isinstance(val, datetime.datetime):
            return val.date().isoformat()
        return val.isoformat()
    return None

def find_header_row(ws):
    """Return 0-based index of row containing 'USER NAME'."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(str(c).strip().upper() == 'USER NAME' for c in row if c is not None):
            return i
    return None

def get_col(headers, *names):
    """Return 0-based column index for first matching header name."""
    upper_names = [n.upper() for n in names]
    for i, h in enumerate(headers):
        if h is not None and str(h).strip().upper() in upper_names:
            return i
    return None

def clean_cnic(raw):
    if not raw:
        return None
    s = str(raw).strip()
    if re.match(r'^\d{5}-\d{7}-\d$', s):
        return s
    digits = re.sub(r'\D', '', s)
    if len(digits) == 13:
        return f'{digits[:5]}-{digits[5:12]}-{digits[12]}'
    return s if s else None

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    env = load_env()
    url = env.get('NEXT_PUBLIC_SUPABASE_URL', '').strip()
    key = env.get('SUPABASE_SERVICE_KEY', '').strip()

    if not url:
        print('ERROR: NEXT_PUBLIC_SUPABASE_URL missing from .env.local')
        return
    if not key:
        key = input('Paste your Supabase service_role key: ').strip()
    if not key:
        print('ERROR: service_role key required')
        return

    sb = create_client(url, key)

    # ── Areas ────────────────────────────────────────────────────────────────
    print('\n[1/4] Inserting areas...')
    all_areas = list(RAHWALI_AREAS.values()) + list(GARRISON_AREAS.values())
    seen_codes = set()
    unique_areas = []
    for a in all_areas:
        if a['code'] not in seen_codes:
            seen_codes.add(a['code'])
            unique_areas.append(a)

    sb.table('areas').upsert(unique_areas, on_conflict='code').execute()
    print(f'  {len(unique_areas)} areas inserted.')

    area_rows = sb.table('areas').select('id,code').execute().data
    area_map = {r['code']: r['id'] for r in area_rows}

    # ── Packages ─────────────────────────────────────────────────────────────
    print('\n[2/4] Collecting packages...')
    pkg_names = set()

    def collect_pkgs(filepath, area_defs):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sname in wb.sheetnames:
            if sname not in area_defs:
                continue
            ws = wb[sname]
            hi = find_header_row(ws)
            if hi is None:
                continue
            headers = list(ws.iter_rows(values_only=True))[hi]
            pkg_col = get_col(headers, 'PKG')
            if pkg_col is None:
                continue
            for row in list(ws.iter_rows(values_only=True))[hi + 1:]:
                pkg = normalize_package(row[pkg_col])
                if pkg:
                    pkg_names.add(pkg)

    collect_pkgs(RAHWALI_FILE, RAHWALI_AREAS)
    collect_pkgs(GARRISON_FILE, GARRISON_AREAS)

    def speed_from(name):
        m = re.search(r'(\d+)', name)
        return int(m.group(1)) if m else 0

    pkg_records = [{'name': n, 'speed_mbps': speed_from(n)} for n in sorted(pkg_names)]
    sb.table('packages').upsert(pkg_records, on_conflict='name').execute()
    print(f'  {len(pkg_records)} packages inserted.')

    pkg_rows = sb.table('packages').select('id,name').execute().data
    pkg_map = {r['name']: r['id'] for r in pkg_rows}

    # ── Customers ────────────────────────────────────────────────────────────
    print('\n[3/4] Inserting customers...')
    total = 0
    skipped = 0

    def insert_sheet(filepath, area_defs, is_garrison):
        nonlocal total, skipped
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sname in wb.sheetnames:
            if sname not in area_defs:
                continue
            area_info = area_defs[sname]
            area_id = area_map.get(area_info['code'])
            ws = wb[sname]
            hi = find_header_row(ws)
            if hi is None:
                print(f'  [{sname}] No header row found — skipping')
                continue
            all_rows = list(ws.iter_rows(values_only=True))
            headers = all_rows[hi]

            col_username = get_col(headers, 'USER NAME')
            col_pkg      = get_col(headers, 'PKG')
            col_iptv     = get_col(headers, 'IPTV')
            col_name     = get_col(headers, 'NAME')
            col_cnic     = get_col(headers, 'CNIC NO', 'CNIC')
            col_phone    = get_col(headers, 'MOBILE NO', 'CELL NO', 'MOB NO', 'MOBILE')
            col_date     = get_col(headers, 'DATE', 'DATE ')
            col_due      = get_col(headers, 'DUE')
            col_remarks  = get_col(headers, 'REMARKS')
            col_onu      = get_col(headers, 'ONU') if is_garrison else None

            # Address column logic
            if sname == 'AIT':
                col_addr = get_col(headers, 'NEW ID NO')
                addr_type = 'id_number'
            elif is_garrison and sname != 'GT ROAD':
                col_addr = get_col(headers, 'ADRESS', 'ADDRESS')
                addr_type = 'text'
            else:
                col_addr = get_col(headers, 'ID NO', 'ID')
                addr_type = 'id_number'

            batch = []
            for row in all_rows[hi + 1:]:
                name_val = row[col_name] if col_name is not None else None
                if not name_val or str(name_val).strip() in ('', 'NAME'):
                    skipped += 1
                    continue

                username_val = None
                if col_username is not None and row[col_username]:
                    u = str(row[col_username]).strip()
                    username_val = u if u else None

                pkg_name = normalize_package(row[col_pkg]) if col_pkg is not None else None
                pkg_id   = pkg_map.get(pkg_name) if pkg_name else None

                iptv_val = row[col_iptv] if col_iptv is not None else None
                has_iptv = iptv_val is not None

                addr_val = None
                if col_addr is not None and row[col_addr]:
                    addr_val = str(row[col_addr]).strip() or None

                due_raw = row[col_due] if col_due is not None else None
                status, due_amount = parse_due(due_raw)

                batch.append({
                    'username':        username_val,
                    'full_name':       str(name_val).strip(),
                    'cnic':            clean_cnic(row[col_cnic]) if col_cnic is not None else None,
                    'phone':           str(row[col_phone]).strip() if col_phone is not None and row[col_phone] else None,
                    'package_id':      pkg_id,
                    'iptv':            has_iptv,
                    'address_type':    addr_type,
                    'address_value':   addr_val,
                    'area_id':         area_id,
                    'connection_date': parse_date(row[col_date]) if col_date is not None else None,
                    'due_amount':      due_amount,
                    'onu_number':      str(row[col_onu]).strip() if col_onu is not None and row[col_onu] else None,
                    'status':          status,
                    'remarks':         str(row[col_remarks]).strip() if col_remarks is not None and row[col_remarks] else None,
                })

            # Insert in chunks of 100
            inserted = 0
            for i in range(0, len(batch), 100):
                chunk = batch[i:i + 100]
                try:
                    sb.table('customers').insert(chunk).execute()
                    inserted += len(chunk)
                except Exception as e:
                    print(f'  [{sname}] Chunk error at {i}: {e}')
            total += inserted
            print(f'  [{sname}] {inserted} records')

    insert_sheet(RAHWALI_FILE,  RAHWALI_AREAS,  is_garrison=False)
    insert_sheet(GARRISON_FILE, GARRISON_AREAS, is_garrison=True)
    print(f'\n  Total customers: {total} inserted, {skipped} skipped (empty rows)')

    # ── Staff seed ───────────────────────────────────────────────────────────
    print('\n[4/4] Seeding staff...')
    staff_seed = [
        {'full_name': 'Bilal Ahmed',  'role': 'technician',     'phone': '+92 301 1128473', 'area_id': area_map.get('AR'),    'is_active': True},
        {'full_name': 'Hassan Raza',  'role': 'recovery_agent', 'phone': '+92 333 7742891', 'area_id': area_map.get('DEF-1'), 'is_active': True},
        {'full_name': 'Fatima Noor',  'role': 'technician',     'phone': '+92 321 3392847', 'area_id': area_map.get('ASK-1'), 'is_active': True},
        {'full_name': 'Ahmed Sheikh', 'role': 'recovery_agent', 'phone': '+92 345 1102984', 'area_id': area_map.get('BZR'),   'is_active': True},
        {'full_name': 'Usman Khan',   'role': 'recovery_agent', 'phone': '+92 302 2297413', 'area_id': area_map.get('AR'),    'is_active': True},
        {'full_name': 'Sara Javed',   'role': 'technician',     'phone': '+92 312 6629187', 'area_id': area_map.get('DEF-2'), 'is_active': True},
        {'full_name': 'Kamran Butt',  'role': 'recovery_agent', 'phone': '+92 321 4419287', 'area_id': area_map.get('ASK-2'), 'is_active': True},
        {'full_name': 'Zainab Malik', 'role': 'technician',     'phone': '+92 300 8847162', 'area_id': area_map.get('AIT'),   'is_active': False},
    ]
    sb.table('staff').insert(staff_seed).execute()
    print(f'  {len(staff_seed)} staff inserted.')
    print('\nMigration complete!')

if __name__ == '__main__':
    main()
