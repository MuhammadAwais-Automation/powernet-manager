"""
Import Bazar + Garrison cable/internet recovery Excel registers into Supabase.

Files:
  - Bazar Area JUNE 2026.xlsx  -> area BZR (Bazaar)
  - Garrison Area JUNE 2026.xlsx -> ASK1, ASK2, DEF1, DEF2 sheets

Run:
  python scripts/import_june_2026_cable_excel.py

Requires NEXT_PUBLIC_SUPABASE_URL + SUPABASE_SERVICE_KEY in .env.local
"""
from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass
from typing import Any

import openpyxl
from supabase import create_client

BAZAR_FILE = r"c:\Users\PC\Downloads\Bazar Area JUNE 2026.xlsx"
GARRISON_FILE = r"c:\Users\PC\Downloads\Garrison Area JUNE 2026.xlsx"

BILLING_MONTH_DEFAULT = "2026-06"
DEF2_BILLING_MONTH = "2026-05"  # sheet title says MAY 2026

AREA_BY_SHEET = {
    "__bazar__": {"code": "BZR", "name": "Bazaar", "type": "garrison"},
    "ASK1": {"code": "ASK-1", "name": "Ask Sector 1", "type": "garrison"},
    "ASK2": {"code": "ASK-2", "name": "Ask Sector 2", "type": "garrison"},
    "DEF1": {"code": "DEF-1", "name": "Defence Sector 1", "type": "garrison"},
    "DEF2": {"code": "DEF-2", "name": "Defence Sector 2", "type": "garrison"},
}

# Bazar layout (1-based columns)
BAZAR_COL = {
    "sno": 1,
    "name": 2,
    "card": 3,
    "address": 4,
    "status": 5,
    "due": 6,
    "rec": 7,
    "bal": 8,
    "phone": 9,
    "remarks": 10,
}

# Garrison layout (1-based columns)
GARRISON_COL = {
    "sno": 1,
    "name": 2,
    "hno": 3,
    "card": 4,
    "status": 5,
    "due": 39,
    "rec": 40,
    "bal": 41,
    "phone_new": 42,
    "phone": 43,
    "remarks": 44,
}


@dataclass
class ParsedRow:
    sheet_key: str
    area_code: str
    billing_month: str
    full_name: str
    card_no: str | None
    address: str | None
    sheet_status: str | None
    due_raw: Any
    rec_raw: Any
    bal_raw: Any
    phone: str | None
    remarks: str | None


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    env_file = os.path.join(root, ".env.local")
    if os.path.exists(env_file):
        with open(env_file, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    env[k.strip()] = v.strip()
    return env


def format_name(raw: str) -> str:
    s = re.sub(r"\s+", " ", raw.strip())
    if not s:
        return s
    # Title case but keep short tokens like "H NO", "SHOP"
    parts = []
    for word in s.split():
        upper = word.upper()
        if upper in ("H", "NO", "SHOP", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z") or re.fullmatch(r"\d+[A-Z]?", upper):
            parts.append(upper)
        elif len(word) <= 2 and word.isalpha():
            parts.append(upper)
        else:
            parts.append(word.capitalize())
    return " ".join(parts)


def format_address(raw: str | None) -> str | None:
    if not raw:
        return None
    s = re.sub(r"\s+", " ", str(raw).strip())
    return s.upper() if s else None


def format_phone_pk(raw: Any) -> str | None:
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw).strip())
    if not digits:
        return None
    if digits.startswith("92") and len(digits) >= 12:
        digits = "0" + digits[2:]
    if len(digits) == 10 and digits.startswith("3"):
        digits = "0" + digits
    if len(digits) == 11 and digits.startswith("03"):
        return f"{digits[:4]}-{digits[4:]}"
    if len(digits) >= 7:
        return digits
    return str(raw).strip() or None


def clean_phone(raw: Any) -> str | None:
    return format_phone_pk(raw)


def clean_card(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        if float(raw).is_integer():
            return str(int(raw))
        return str(raw).strip()
    s = str(raw).strip()
    return s or None


MAX_CABLE_BILL = 50_000  # PKR — reject phone numbers / bad Excel cells


def sanitize_bill_amount(amount: int | None) -> int | None:
    if amount is None or amount <= 0:
        return None
    if amount > MAX_CABLE_BILL:
        return None
    return amount


def parse_due_cell(
    due_val: Any,
    sheet_status: str | None,
) -> tuple[str, int | None, bool, bool]:
    """
    Returns: customer_status, bill_amount, has_cable, has_internet
    """
    sheet_st = (sheet_status or "").strip().upper()
    base_disconnected = sheet_st == "DECO"

    if due_val is None:
        status = "disconnected" if base_disconnected else "active"
        return status, None, True, False

    if isinstance(due_val, (int, float)):
        amount = int(round(float(due_val)))
        if amount < 0:
            amount = None
        status = "disconnected" if base_disconnected and not amount else "active"
        if amount == 0:
            return "free", 0, True, False
        return status, amount, True, False

    token = str(due_val).strip().upper()
    # Phone numbers sometimes land in DUE column — not a bill amount
    if re.match(r"^0?3\d{2}[-\s]?\d{6,7}$", str(due_val).strip()):
        return "active", None, True, False
    if token in ("DC", ""):
        return "disconnected", None, True, False
    if token == "FREE":
        return "free", 0, True, False
    if token in ("NET",):
        return "active", None, False, True
    if token == "STOP":
        return "suspended", None, True, False
    if token == "TDC":
        return "tdc", None, True, False
    if token in ("ADV", "ADVANCE"):
        return "active", None, True, False

    # Text amounts e.g. "NET" variants
    if token.startswith("DC"):
        return "disconnected", None, True, False

    try:
        amount = int(float(token))
        status = "disconnected" if base_disconnected and amount <= 0 else "active"
        return status, amount, True, False
    except ValueError:
        status = "disconnected" if base_disconnected else "active"
        return status, None, True, False


def parse_rec_amount(rec_val: Any) -> int | None:
    if isinstance(rec_val, (int, float)):
        amt = int(round(float(rec_val)))
        return amt if amt > 0 else None
    return None


def parse_bal_amount(bal_val: Any) -> int | None:
    if isinstance(bal_val, (int, float)):
        amt = int(round(float(bal_val)))
        return amt if amt > 0 else None
    return None


def make_customer_code(area_code: str, card: str | None, name: str, row_idx: int) -> str:
    card_slug = re.sub(r"\D", "", card or "")[:10] or "NC"
    name_slug = re.sub(r"[^A-Z0-9]+", "", (name or "X").upper())[:6] or "X"
    return f"CAB-{area_code}-{row_idx:05d}-{card_slug}-{name_slug}"


def cell(ws, row: int, col: int) -> Any:
    if col <= 0:
        return None
    return ws.cell(row, col).value


def iter_bazar_rows(ws) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    for r in range(4, ws.max_row + 1):
        sno = cell(ws, r, BAZAR_COL["sno"])
        if not isinstance(sno, (int, float)):
            continue
        name = cell(ws, r, BAZAR_COL["name"])
        if not name or not str(name).strip():
            continue
        rows.append(
            ParsedRow(
                sheet_key="__bazar__",
                area_code="BZR",
                billing_month=BILLING_MONTH_DEFAULT,
                full_name=str(name).strip(),
                card_no=clean_card(cell(ws, r, BAZAR_COL["card"])),
                address=str(cell(ws, r, BAZAR_COL["address"])).strip()
                if cell(ws, r, BAZAR_COL["address"])
                else None,
                sheet_status=str(cell(ws, r, BAZAR_COL["status"])).strip()
                if cell(ws, r, BAZAR_COL["status"])
                else None,
                due_raw=cell(ws, r, BAZAR_COL["due"]),
                rec_raw=cell(ws, r, BAZAR_COL["rec"]),
                bal_raw=cell(ws, r, BAZAR_COL["bal"]),
                phone=clean_phone(cell(ws, r, BAZAR_COL["phone"])),
                remarks=str(cell(ws, r, BAZAR_COL["remarks"])).strip()
                if cell(ws, r, BAZAR_COL["remarks"])
                else None,
            )
        )
    return rows


def iter_garrison_rows(ws, sheet_name: str) -> list[ParsedRow]:
    area = AREA_BY_SHEET[sheet_name]
    month = DEF2_BILLING_MONTH if sheet_name == "DEF2" else BILLING_MONTH_DEFAULT
    rows: list[ParsedRow] = []
    for r in range(4, ws.max_row + 1):
        sno = cell(ws, r, GARRISON_COL["sno"])
        if not isinstance(sno, (int, float)):
            continue
        name = cell(ws, r, GARRISON_COL["name"])
        if not name or not str(name).strip():
            continue
        hno = cell(ws, r, GARRISON_COL["hno"])
        address = str(hno).strip() if hno else None
        phone = clean_phone(cell(ws, r, GARRISON_COL["phone"])) or clean_phone(
            cell(ws, r, GARRISON_COL["phone_new"])
        )
        rows.append(
            ParsedRow(
                sheet_key=sheet_name,
                area_code=area["code"],
                billing_month=month,
                full_name=str(name).strip(),
                card_no=clean_card(cell(ws, r, GARRISON_COL["card"])),
                address=address,
                sheet_status=str(cell(ws, r, GARRISON_COL["status"])).strip()
                if cell(ws, r, GARRISON_COL["status"])
                else None,
                due_raw=cell(ws, r, GARRISON_COL["due"]),
                rec_raw=cell(ws, r, GARRISON_COL["rec"]),
                bal_raw=cell(ws, r, GARRISON_COL["bal"]),
                phone=phone,
                remarks=str(cell(ws, r, GARRISON_COL["remarks"])).strip()
                if cell(ws, r, GARRISON_COL["remarks"])
                else None,
            )
        )
    return rows


def load_all_rows() -> list[ParsedRow]:
    out: list[ParsedRow] = []
    wb = openpyxl.load_workbook(BAZAR_FILE, data_only=True)
    out.extend(iter_bazar_rows(wb.active))
    wb.close()

    wb = openpyxl.load_workbook(GARRISON_FILE, data_only=True)
    for sheet in ("ASK1", "ASK2", "DEF1", "DEF2"):
        if sheet in wb.sheetnames:
            out.extend(iter_garrison_rows(wb[sheet], sheet))
    wb.close()
    return out


def parse_cable_type(sheet_status: str | None) -> str | None:
    token = (sheet_status or "").strip().upper()
    if token in ("DECO", "DIGITAL", "DC"):
        return "digital"
    if token in ("ANAL", "ANALOG"):
        return "analog"
    return None


def build_customer_record(
    row: ParsedRow,
    area_id: str,
    row_idx: int,
) -> dict[str, Any]:
    status, bill_amount, has_cable, has_internet = parse_due_cell(
        row.due_raw, row.sheet_status
    )
    rec_amount = parse_rec_amount(row.rec_raw)
    bal_amount = parse_bal_amount(row.bal_raw)

    address_value = row.address
    house_id = None  # avoid global unique constraint clashes on shared addresses

    remarks_parts = []
    if row.card_no:
        remarks_parts.append(f"Card: {row.card_no}")
    if row.remarks:
        remarks_parts.append(row.remarks)
    if row.sheet_status:
        remarks_parts.append(f"Sheet status: {row.sheet_status}")
    if bal_amount:
        remarks_parts.append(f"Balance carry: Rs {bal_amount}")
    remarks = " | ".join(remarks_parts) if remarks_parts else None

    due_amount = sanitize_bill_amount(bill_amount)

    return {
        "customer_code": make_customer_code(row.area_code, row.card_no, row.full_name, row_idx),
        "full_name": format_name(row.full_name),
        "phone": row.phone,
        "house_id": house_id,
        "address_type": "text",
        "address_value": format_address(address_value),
        "area_id": area_id,
        "status": status,
        "has_cable": has_cable,
        "has_internet": has_internet,
        "cable_type": parse_cable_type(row.sheet_status),
        "iptv": bool(row.card_no) and has_cable,
        "due_amount": due_amount,
        "remarks": remarks,
        "_bill_amount": sanitize_bill_amount(bill_amount),
        "_rec_amount": rec_amount,
        "_billing_month": row.billing_month,
    }


def match_key(area_id: str, card: str | None, name: str, address: str | None) -> str:
    card_part = card or ""
    addr_part = (address or "").upper().strip()
    name_part = name.upper().strip()
    return f"{area_id}|{card_part}|{name_part}|{addr_part}"


def main() -> int:
    env = load_env()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL", "").strip()
    key = env.get("SUPABASE_SERVICE_KEY", "").strip()
    if not url or not key:
        print("ERROR: NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_KEY required in .env.local")
        return 1

    if not os.path.exists(BAZAR_FILE):
        print(f"ERROR: Missing file: {BAZAR_FILE}")
        return 1
    if not os.path.exists(GARRISON_FILE):
        print(f"ERROR: Missing file: {GARRISON_FILE}")
        return 1

    rows = load_all_rows()
    print(f"Parsed {len(rows)} subscriber rows from Excel")

    # Deduplicate identical area+card+name+address (keep last row)
    deduped: dict[str, ParsedRow] = {}
    for i, row in enumerate(rows, start=1):
        area_id_preview = row.area_code
        k = f"{area_id_preview}|{row.card_no or ''}|{row.full_name.upper()}|{(row.address or '').upper()}"
        deduped[k] = row
    rows = list(deduped.values())
    print(f"After dedupe: {len(rows)} unique subscriber rows")

    sb = create_client(url, key)

    # Ensure areas
    area_defs = {v["code"]: v for v in AREA_BY_SHEET.values()}
    sb.table("areas").upsert(list(area_defs.values()), on_conflict="code").execute()
    area_rows = sb.table("areas").select("id,code").execute().data or []
    area_map = {r["code"]: r["id"] for r in area_rows}

    # Load existing customers for matching (paginated)
    existing: list[dict] = []
    offset = 0
    page_size = 1000
    while True:
        chunk = (
            sb.table("customers")
            .select("id,customer_code,full_name,house_id,address_value,area_id,has_cable,has_internet,remarks")
            .range(offset, offset + page_size - 1)
            .execute()
            .data
            or []
        )
        existing.extend(chunk)
        if len(chunk) < page_size:
            break
        offset += page_size
    print(f"Loaded {len(existing)} existing customers from DB")
    existing_by_key: dict[str, dict] = {}
    existing_by_card: dict[str, dict] = {}
    existing_by_code: dict[str, dict] = {}
    for c in existing:
        k = match_key(
            c["area_id"] or "",
            c.get("house_id"),
            c.get("full_name") or "",
            c.get("address_value"),
        )
        existing_by_key[k] = c
        if c.get("customer_code", "").startswith("CAB-"):
            existing_by_code[c["customer_code"]] = c
        card_m = re.search(r"Card:\s*(\S+)", c.get("remarks") or "")
        if card_m and c.get("area_id"):
            existing_by_card[f"{c['area_id']}|{card_m.group(1)}"] = c

    to_insert: list[dict[str, Any]] = []
    to_update: list[tuple[str, dict[str, Any]]] = []
    bill_ops: list[dict[str, Any]] = []
    internet_markers: list[str] = []

    for idx, row in enumerate(rows, start=1):
        area_id = area_map.get(row.area_code)
        if not area_id:
            print(f"WARN: Unknown area {row.area_code}")
            continue

        built = build_customer_record(row, area_id, idx)
        bill_amount = built.pop("_bill_amount")
        rec_amount = built.pop("_rec_amount")
        billing_month = built.pop("_billing_month")

        key = match_key(area_id, built.get("house_id"), built["full_name"], built.get("address_value"))
        card_key = f"{area_id}|{row.card_no}" if row.card_no else None
        existing_rec = (
            existing_by_key.get(key)
            or (existing_by_card.get(card_key) if card_key else None)
            or existing_by_code.get(built["customer_code"])
        )
        if existing_rec:
            cid = existing_rec["id"]
            patch = {k: v for k, v in built.items() if k != "customer_code"}
            # Preserve dual-service flags when updating
            if existing_rec.get("has_internet") and not patch["has_internet"]:
                patch["has_internet"] = True
            if existing_rec.get("has_cable") and not patch["has_cable"]:
                patch["has_cable"] = True
            to_update.append((cid, patch))
            customer_id = cid
        else:
            to_insert.append(built)
            customer_id = None

        bill_ops.append(
            {
                "customer_id": customer_id,
                "match_key": key,
                "card_key": card_key,
                "customer_code": built["customer_code"],
                "billing_month": billing_month,
                "bill_amount": bill_amount,
                "rec_amount": rec_amount,
                "has_cable": built["has_cable"],
                "has_internet": built["has_internet"],
            }
        )
        if built["has_internet"] and not built["has_cable"]:
            internet_markers.append(key)

    print(f"Insert {len(to_insert)} | Update {len(to_update)} customers")

    # Insert new customers in chunks
    inserted_map: dict[str, str] = {}
    for i in range(0, len(to_insert), 50):
        chunk = to_insert[i : i + 50]
        try:
            res = sb.table("customers").insert(chunk).select(
                "id,customer_code,area_id,house_id,full_name,address_value"
            ).execute()
        except Exception as e:
            print(f"Chunk insert error at {i}: {e}")
            for rec in chunk:
                try:
                    one = sb.table("customers").insert(rec).select(
                        "id,customer_code,area_id,house_id,full_name,address_value"
                    ).execute()
                    if one.data:
                        for row in one.data:
                            k = match_key(
                                row["area_id"],
                                row.get("house_id"),
                                row.get("full_name") or "",
                                row.get("address_value"),
                            )
                            inserted_map[k] = row["id"]
                except Exception as e2:
                    print(f"  Skip {rec.get('full_name')}: {e2}")
            continue
        for rec in res.data or []:
            k = match_key(
                rec["area_id"],
                rec.get("house_id"),
                rec.get("full_name") or "",
                rec.get("address_value"),
            )
            inserted_map[k] = rec["id"]

    # Updates
    updated = 0
    for cid, patch in to_update:
        try:
            sb.table("customers").update(patch).eq("id", cid).execute()
            updated += 1
        except Exception as e:
            if "house_id_uidx" in str(e):
                patch2 = dict(patch)
                patch2["house_id"] = None
                sb.table("customers").update(patch2).eq("id", cid).execute()
                updated += 1
            else:
                print(f"  Update failed {cid}: {e}")

    # Resolve customer ids for bills
    all_ids = {**{k: c["id"] for k, c in existing_by_key.items()}, **inserted_map}
    for c in existing:
        if c.get("customer_code", "").startswith("CAB-"):
            all_ids[f"code|{c['customer_code']}"] = c["id"]
    for op in bill_ops:
        if not op["customer_id"]:
            op["customer_id"] = (
                all_ids.get(op["match_key"])
                or all_ids.get(f"code|{op['customer_code']}")
                or (existing_by_card.get(op["card_key"]) or {}).get("id")
                if op.get("card_key")
                else None
            )

    # Cable bills for June (and May DEF2)
    cable_created = 0
    cable_paid = 0
    cable_skipped = 0
    for op in bill_ops:
        cid = op["customer_id"]
        if not cid or not op["has_cable"]:
            continue
        amount = sanitize_bill_amount(op["bill_amount"])
        if amount is None:
            if op["bill_amount"] and op["bill_amount"] > MAX_CABLE_BILL:
                cable_skipped += 1
            continue
        month = op["billing_month"]
        rec = sanitize_bill_amount(op["rec_amount"]) or 0
        paid = min(rec, amount) if rec else 0
        status = "paid" if paid >= amount else "pending"
        bill_row = {
            "customer_id": cid,
            "amount": amount,
            "paid_amount": paid,
            "month": month,
            "status": status,
            "payment_source": "manual",
        }
        if paid > 0:
            bill_row["payment_method"] = "cash"
            bill_row["payment_source"] = "office"
        try:
            sb.table("cable_bills").upsert(
                bill_row,
                on_conflict="customer_id,month",
            ).execute()
            cable_created += 1
            if paid > 0:
                cable_paid += 1
        except Exception as e:
            if "out of range" in str(e) or "22003" in str(e):
                cable_skipped += 1
                continue
            try:
                existing_bill = (
                    sb.table("cable_bills")
                    .select("id")
                    .eq("customer_id", cid)
                    .eq("month", month)
                    .limit(1)
                    .execute()
                )
                rows_found = existing_bill.data or []
                if rows_found:
                    sb.table("cable_bills").update(bill_row).eq("id", rows_found[0]["id"]).execute()
                    cable_created += 1
                else:
                    sb.table("cable_bills").insert(bill_row).execute()
                    cable_created += 1
            except Exception as e2:
                print(f"  Cable bill skip {cid} {month}: {e2}")
                cable_skipped += 1

    # Internet-only rows: ensure has_internet flag (already on customer)
    print(f"Updated {updated} existing customers")
    print(f"Inserted {len(inserted_map)} new customers")
    print(f"Cable bills upserted: {cable_created} ({cable_paid} with payment recorded, {cable_skipped} skipped bad amounts)")
    print(f"Internet-only (NET) subscribers: {len(internet_markers)}")
    print("Import complete.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
